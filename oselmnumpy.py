# coding=utf-8
# https://github.com/numenta/htmresearch/blob/master/htmresearch/algorithms/online_extreme_learning_machine.py
# ----------------------------------------------------------------------
# Numenta Platform for Intelligent Computing (NuPIC)
# Copyright (C) 2015, Numenta, Inc.  Unless you have an agreement
# with Numenta, Inc., for a separate license for this software code, the
# following terms and conditions apply:
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero Public License version 3 as
# published by the Free Software Foundation.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
# See the GNU Affero Public License for more details.
#
# You should have received a copy of the GNU Affero Public License
# along with this program.  If not, see http://www.gnu.org/licenses.
#
# http://numenta.org/licenses/
# ----------------------------------------------------------------------
#import xlrd
import openpyxl
import numpy as np
import os
from numpy.linalg import pinv

"""
Implementation of the online-sequential extreme learning machine
Reference:
N.-Y. Liang, G.-B. Huang, P. Saratchandran, and N. Sundararajan,
“A Fast and Accurate On-line Sequential Learning Algorithm for Feedforward
Networks," IEEE Transactions on Neural Networks, vol. 17, no. 6, pp. 1411-1423
"""


class OSELM(object):
    # def __init__(self, inputs, outputs, numHiddenNeurons, activationFunction):
    def __init__(self):

        #file_loc = "input_data.xlsx"
        #wkb = xlrd.open_workbook(input_data.xlsx)
        #sheet = wkb.sheet_by_index(0)

        #openpyxl can read xlsx format, xlrd can't
        wb = openpyxl.load_workbook('input_data.xlsx')
        sheet = wb.active

        #this is all used for reading info from the excel spreadsheet into an array named data
        firstRow = 2
        firstCol = 2
        nCols = sheet.max_column
        nRows = sheet.max_row

        data = np.array([[cell.value for cell in row] for row in sheet.iter_rows()])
        matrix = data[(firstRow - 1):(firstRow - 1 + nRows), (firstCol - 1):(firstCol - 1 + nCols)]
        #print(data)

        ##Original code below
        #Did not work on latest pycharm and python, fixed by changing to openpyxl
        #for row in range(1, sheet.max_row):
            #for col in range(1, sheet.max_column):
                #matrix.append(row)#(row, col))
        #self.matrix = np.reshape(matrix, (sheet.max_row, sheet.max_column))

        #print(self.matrix)

        self.activationFunction = "sig"

        self.inputs_matrix = matrix[:, 0:4]

        self.inputs = self.inputs_matrix.shape[1]

        print(self.inputs)
        self.outputs_matrix = matrix[:, 4:6]
        self.outputs = self.outputs_matrix.shape[1]

        self.numHiddenNeurons = 20

        # input to hidden weights
        self.inputWeights = np.random.random((self.numHiddenNeurons, self.inputs))

        # bias of hidden units
        self.bias = np.random.random((1, self.numHiddenNeurons)) * 2 - 1

        # hidden to output layer connection
        self.beta = np.random.random((self.numHiddenNeurons, self.outputs))

        # auxiliary matrix used for sequential learning
        self.M = None
        self.initializePhase(self.inputs_matrix, self.outputs_matrix)
        self.train(self.inputs_matrix, self.outputs_matrix)
        os.system("./recordRSSI")
        # fname = input("file_mat.txt")
        text_online = []
        # count=0
        with open('file_mat.txt', 'r') as f:
            for line in f:
                words = line.split()

                for i in words:
                    r = i.strip('-')
                    t = r.strip('.00')

                    if (t.isdigit()):
                        print('essss', t)
                        t = float(t)
                        text_online.append(-1 * t)

        f.close()
        text_online = np.matrix(text_online)
        print('test RSSI', text_online)
        self.predict(text_online)

    def sigmoidActFunc(self, features, weights, bias):
        assert (features.shape[1] == weights.shape[1])
        (numSamples, numInputs) = features.shape
        (numHiddenNeuron, numInputs) = weights.shape
        V = np.dot(features, np.transpose(weights))
        for i in range(numHiddenNeuron):
            V[:, i] += bias[0, i]
        H = 1 / (1 + np.exp(-V))
        #print('dddd')
        return H

    def calculateHiddenLayerActivation(self, features):
        """
        Calculate activation level of the hidden layer
        :param features feature matrix with dimension (numSamples, numInputs)
        :return: activation level (numSamples, numHiddenNeurons)
        """
        if self.activationFunction is "sig":
            H = self.sigmoidActFunc(features, self.inputWeights, self.bias)
        else:
            print(" Unknown activation function type")
            raise NotImplementedError
        return H

    def initializePhase(self, features, targets):

        """
        Step 1: Initialization phase
        :param features feature matrix with dimension (numSamples, numInputs)
        :param targets target matrix with dimension (numSamples, numOutputs)
        """

        assert features.shape[0] == targets.shape[0]
        assert features.shape[1] == self.inputs
        assert targets.shape[1] == self.outputs
        #print('fff', targets)

        # randomly initialize the input->hidden connections
        self.inputWeights = np.random.random((self.numHiddenNeurons, self.inputs))
        self.inputWeights = self.inputWeights * 2 - 1

        if self.activationFunction is "sig":
            self.bias = np.random.random((1, self.numHiddenNeurons)) * 2 - 1
        else:
            print(" Unknown activation function type")
            raise NotImplementedError

        H0 = self.calculateHiddenLayerActivation(features)
        self.M = pinv(np.dot(np.transpose(H0), H0))
        self.beta = np.dot(pinv(H0), targets)
        print('beta', self.beta)

    def train(self, features, targets):

        """
        Step 2: Sequential learning phase
        :param features feature matrix with dimension (numSamples, numInputs)
        :param targets target matrix with dimension (numSamples, numOutputs)
        """

        (numSamples, numOutputs) = targets.shape
        assert features.shape[0] == targets.shape[0]

        H = self.calculateHiddenLayerActivation(features)
        Ht = np.transpose(H)
        try:
            self.M -= np.dot(self.M,
                             np.dot(Ht, np.dot(
                                 pinv(np.eye(numSamples) + np.dot(H, np.dot(self.M, Ht))),
                                 np.dot(H, self.M))))

            self.beta += np.dot(self.M, np.dot(Ht, targets - np.dot(H, self.beta)))
            print('beta2', self.beta)
        except np.linalg.linalg.LinAlgError:
            print("SVD not converge, ignore the current training cycle")
        # else:
        #   raise RuntimeError

    def predict(self, features):
        """
        Make prediction with feature matrix
        :param features: feature matrix with dimension (numSamples, numInputs)
        :return: predictions with dimension (numSamples, numOutputs)
        """
        H = self.calculateHiddenLayerActivation(features)
        prediction = np.dot(H, self.beta)
        print('hiiii', prediction)
        return prediction


OSELM()
